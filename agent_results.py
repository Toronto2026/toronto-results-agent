# -*- coding: utf-8 -*-
from __future__ import annotations
"""
Агент публікації результатів фестивалю TORONTO
===============================================
Збирає оцінки від членів журі → генерує PDF таблицю результатів → записує Laureate у Bitrix24

Використання:
    python agent_results.py --folder "КВІТЕНЬ 2026/Журі оцінки" --month "Квітень 2026"
    python agent_results.py --folder "..." --month "..." --bitrix-url "https://..." --write-bitrix

Залежності:
    pip install openpyxl reportlab requests
"""

import argparse
import os
import sys
import re
from collections import defaultdict

import openpyxl
import requests
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


# ---------------------------------------------------------------------------
# Шрифти (DejaVu — підтримує кирилицю, завантажується разом з reportlab)
# ---------------------------------------------------------------------------
def _register_fonts():
    """Реєструє Arial (Windows) або DejaVu — обидва підтримують кирилицю."""
    # Пробуємо Arial (є на всіх Windows)
    win_fonts = r"C:\Windows\Fonts"
    arial_r = os.path.join(win_fonts, "arial.ttf")
    arial_b = os.path.join(win_fonts, "arialbd.ttf")
    if os.path.exists(arial_r) and os.path.exists(arial_b):
        try:
            pdfmetrics.registerFont(TTFont("Arial",      arial_r))
            pdfmetrics.registerFont(TTFont("Arial-Bold", arial_b))
            return "Arial", "Arial-Bold"
        except Exception:
            pass
    # Пробуємо DejaVu — перевіряємо кілька відомих розташувань (Windows, Linux, поруч зі скриптом)
    dejavu_search = [
        os.path.dirname(__file__),
        win_fonts,
        "/usr/share/fonts/truetype/dejavu",          # Debian/Ubuntu/Streamlit Cloud
        "/usr/share/fonts/dejavu",
        "/usr/share/fonts/TTF",
    ]
    for folder in dejavu_search:
        r = os.path.join(folder, "DejaVuSans.ttf")
        b = os.path.join(folder, "DejaVuSans-Bold.ttf")
        if os.path.exists(r) and os.path.exists(b):
            try:
                pdfmetrics.registerFont(TTFont("DejaVu",      r))
                pdfmetrics.registerFont(TTFont("DejaVu-Bold", b))
                return "DejaVu", "DejaVu-Bold"
            except Exception:
                pass
    return "Helvetica", "Helvetica-Bold"


FONT_REGULAR, FONT_BOLD = _register_fonts()


# ---------------------------------------------------------------------------
# Конвертація значення Laureate
# ---------------------------------------------------------------------------
def convert_laureate(value) -> str:
    """
    Конвертує будь-який формат оцінки у фінальний рядок.
    Підтримує: 1/2/3, I/II/III, 1 місце, 1st/2nd/3d degree,
               Gran Pri, Гран Прі, нема, не відкрила, None → 1st degree
    """
    if value is None:
        return "1st degree"
    v = str(value).strip()
    if v == "":
        return "1st degree"
    low = v.lower()

    # Гран Прі — перевіряємо першим (з тире, без тире, різні варіанти)
    low_nodash = low.replace("-", " ").replace("'", " ")
    if any(k in low_nodash for k in ["gran pri", "гран прі", "гран при", "gran prize"]):
        return "Gran Pri"

    # Числові та рядкові варіанти 1/2/3
    # Витягуємо першу цифру або римські
    import re

    # Спочатку — точні відповідності
    exact = {
        "1": "1st degree", "1st": "1st degree", "1st degree": "1st degree",
        "1 місце": "1st degree", "1 место": "1st degree", "перше": "1st degree",
        "перший": "1st degree", "першe": "1st degree", "i": "1st degree",
        "2": "2nd degree", "2nd": "2nd degree", "2nd degree": "2nd degree",
        "2 місце": "2nd degree", "2 место": "2nd degree", "друге": "2nd degree",
        "другий": "2nd degree", "ii": "2nd degree",
        "3": "3d degree",  "3d": "3d degree",  "3d degree": "3d degree",
        "3rd": "3d degree", "3rd degree": "3d degree",
        "3 місце": "3d degree", "3 место": "3d degree", "третє": "3d degree",
        "третій": "3d degree", "iii": "3d degree",
    }
    if low in exact:
        return exact[low]

    # Числа з текстом: "1 ступінь", "2 ступень" тощо
    m = re.match(r'^(\d)', v)
    if m:
        n = int(m.group(1))
        if n == 1: return "1st degree"
        if n == 2: return "2nd degree"
        if n == 3: return "3d degree"

    # Римські цифри — шукаємо ІІІ/ІІ/І в будь-якому місці рядка
    # Лауреат ІІІ ступеня, Лауреат ІІ ступеня, Лауреат І ступеня
    if re.search(r'\biii\b', low) or re.search(r'ііі', low): return "3d degree"
    if re.search(r'\bii\b',  low) or re.search(r'іі',  low): return "2nd degree"
    if re.search(r'\bi\b',   low) or re.search(r'\bі\b', low): return "1st degree"

    # Дипломант → 3d degree
    if "дипломант" in low:
        return "3d degree"

    # Все інше (нема, не відкрила, error, тощо) → 1st degree
    return "1st degree"


# ---------------------------------------------------------------------------
# Читання файлів журі
# ---------------------------------------------------------------------------
SKIP_PIB_PREFIXES = ("сума", "гонорар", "разом", "total", "підсумок")

def _is_data_row(row_id, pib) -> bool:
    """Повертає True якщо рядок є учасником, а не технічним."""
    if pib is None:
        return False
    pib_s = str(pib).strip().lower()
    for prefix in SKIP_PIB_PREFIXES:
        if pib_s.startswith(prefix):
            return False
    return True


_COUNTRY_KEYWORDS = {
    "Молдова":  ["moldav", "moldova", "moldovei", "кишинів", "kichinev", "chișinău",
                 "молдов", "молдав", "kišinev"],
    "Литва":    ["klaipėd", "vilnius", "lietuv", "kauna", "литв", "литов", "паневеж"],
    "Польща":   ["poland", "warszawa", "kraków", "krakow", "польщ", "варшав", "польськ"],
    "Германія": ["german", "berlin", "münchen", "hamburg", "нiмеч", "берлін"],
    "Франція":  ["france", "paris", "lyon", "франц", "париж"],
    "Ізраїль":  ["israel", "tel aviv", "ізраїл", "ізраілю"],
    "Канада":   ["canada", "toronto", "montreal", "канад"],
    "США":      ["usa", "united states", "new york", "сша", "америк"],
}

def detect_country(school: str) -> str:
    """Визначає країну за назвою навчального закладу."""
    sl = (school or "").lower()
    for country, keywords in _COUNTRY_KEYWORDS.items():
        if any(kw in sl for kw in keywords):
            return country
    return "Україна"


def read_jury_file(path: str) -> tuple[list[dict], list[str]]:
    """
    Читає один xlsx файл журі.
    Повертає (список рядків, лог-повідомлення).
    """
    fname = os.path.basename(path)
    log = []
    log.append(f"📂 Файл: {fname}")

    wb = openpyxl.load_workbook(path)
    results = []

    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
        except Exception as e:
            log.append(f"  ⚠️ Аркуш '{sheet_name}': не вдалося відкрити ({e})")
            continue
        if not hasattr(ws, "iter_rows"):
            log.append(f"  ⏭️ Аркуш '{sheet_name}': діаграма, пропущено")
            continue

        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        log.append(f"  📋 Аркуш '{sheet_name}' | рядків: {ws.max_row - 1}")
        log.append(f"     Колонки: {headers}")

        def col(name):
            try:
                return headers.index(name)
            except ValueError:
                return None

        idx_id   = col("ID")
        idx_pib  = col("ПІБ Учасника")
        idx_nom  = col("Номінація")
        idx_vik  = col("Вікова категорія")
        idx_nazv = col("Назва або опис роботи") or col("Назва роботи")
        idx_sch  = col("Назва навчального закладу")
        # Laureate або Ступінь (файли ДПМ)
        idx_lau  = col("Laureate") or col("Ступінь") or col("Оцінка") or col("Місце")
        idx_kom  = col("Коментар Журі") or col("Коментар журі")

        # Лог знайдених/відсутніх колонок
        col_status = []
        for cname, cidx in [("ID", idx_id), ("ПІБ Учасника", idx_pib),
                             ("Номінація", idx_nom), ("Назва або опис роботи", idx_nazv),
                             ("Laureate", idx_lau)]:
            if cidx is None:
                col_status.append(f"❌ '{cname}' — НЕ ЗНАЙДЕНО")
            else:
                col_status.append(f"✅ '{cname}' → [{cidx}]")
        for s in col_status:
            log.append(f"     {s}")

        if idx_pib is None or idx_lau is None:
            log.append(f"  🚫 Аркуш пропущено: немає обов'язкових колонок 'ПІБ Учасника' або 'Laureate'")
            continue

        no_id_count = 0
        no_score_count = 0
        no_nazva_count = 0
        sheet_rows = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            pib = row[idx_pib] if idx_pib is not None else None
            rid = row[idx_id]  if idx_id  is not None else None

            if not _is_data_row(rid, pib):
                continue

            sheet_rows += 1
            lau_raw = row[idx_lau] if idx_lau is not None else None
            nazva   = row[idx_nazv] if idx_nazv is not None else None

            if rid is None:
                no_id_count += 1
            if lau_raw is None:
                no_score_count += 1
            if not nazva:
                no_nazva_count += 1

            school = str(row[idx_sch]).strip() if idx_sch is not None and row[idx_sch] else ""
            results.append({
                "id":       rid,
                "pib":      str(pib).strip() if pib else "",
                "nom":      str(row[idx_nom]).strip()  if idx_nom  is not None and row[idx_nom]  else "",
                "vik":      str(row[idx_vik]).strip()  if idx_vik  is not None and row[idx_vik]  else "",
                "nazva":    str(nazva).strip() if nazva else "",
                "school":   school,
                "country":  detect_country(school),
                "laureate": convert_laureate(lau_raw),
                "raw_laureate": str(lau_raw) if lau_raw is not None else "None",
                "comment":  str(row[idx_kom]).strip() if idx_kom is not None and row[idx_kom] else "",
                "source":   fname,
            })

        log.append(f"  ✅ Зчитано: {sheet_rows} учасників")
        if no_id_count:
            log.append(f"  ⚠️ Без ID: {no_id_count} — запис у Bitrix24 неможливий")
        if no_score_count:
            log.append(f"  ⚠️ Без оцінки (Laureate=None): {no_score_count} → буде '1st degree'")
        if no_nazva_count:
            log.append(f"  ℹ️ Без назви роботи: {no_nazva_count} — колонка порожня у джерелі")

    log.append(f"  📊 Разом з файлу: {len(results)} рядків")
    return results, log


def read_all_jury_with_log(folder: str) -> tuple[list[dict], list[str]]:
    """Читає всі xlsx файли у папці. Повертає (рядки, повний лог)."""
    all_rows = []
    full_log = []
    for fname in sorted(os.listdir(folder)):
        if fname.lower().endswith(".xlsx") and not fname.startswith("~$"):
            path = os.path.join(folder, fname)
            try:
                rows, log = read_jury_file(path)
                all_rows.extend(rows)
                full_log.extend(log)
                full_log.append("")
                print(f"  {fname}: {len(rows)} рядків")
            except Exception as e:
                msg = f"  ❌ ПОМИЛКА {fname}: {e}"
                full_log.append(msg)
                print(msg)
    return all_rows, full_log


def read_all_jury(folder: str) -> list[dict]:
    """Читає всі xlsx файли у папці (без логу — для CLI)."""
    all_rows = []
    for fname in sorted(os.listdir(folder)):
        if fname.lower().endswith(".xlsx") and not fname.startswith("~$"):
            path = os.path.join(folder, fname)
            try:
                rows, log = read_jury_file(path)
                print(f"  {fname}: {len(rows)} рядків")
                for line in log:
                    print(f"    {line}")
                all_rows.extend(rows)
            except Exception as e:
                print(f"  ПОМИЛКА {fname}: {e}")
    return all_rows


# ---------------------------------------------------------------------------
# Генерація PDF — чистий дизайн "легкий пошук"
# ---------------------------------------------------------------------------
import re as _re

# Кольори рядків по ступеню (пастельні — не втомлюють очі)
ROW_COLORS = {
    "Gran Pri":   colors.HexColor("#FFE082"),  # тепле золото
    "1st degree": colors.HexColor("#DCEDC8"),  # світло-зелений
    "2nd degree": colors.HexColor("#BBDEFB"),  # світло-блакитний
    "3d degree":  colors.HexColor("#E1BEE7"),  # світло-ліловий
}
# Текст в клітинці Laureate (кольоровий жирний)
LAU_TEXT_COLORS = {
    "Gran Pri":   colors.HexColor("#E65100"),  # темно-помаранчевий
    "1st degree": colors.HexColor("#2E7D32"),  # темно-зелений
    "2nd degree": colors.HexColor("#1565C0"),  # темно-синій
    "3d degree":  colors.HexColor("#6A1B9A"),  # темно-ліловий
}

HEADER_TEXT = (
    "Вітаємо всіх учасників фестивалю з чудовими результатами!\n"
    "Як користуватися таблицею?\n"
    "Знаходимо у стовпчику ПІБ Учасника назву колективу або прізвище.\n"
    "Сортування за алфавітом. Навпроти ПІБ учасника ви бачите диплом лауреата від 3 до 1-ї або Гран Прі."
)

def _clean_nazva(text: str) -> str:
    """Прибирає URL з назви роботи — лишає тільки текст."""
    if not text:
        return ""
    # видаляємо http/https посилання
    text = _re.sub(r'https?://\S+', '', text).strip()
    # видаляємо youtu.be/... та www.youtube.com/...
    text = _re.sub(r'(?:youtu\.be|youtube\.com)\S*', '', text).strip()
    return text.strip()


def build_pdf(rows: list[dict], output_path: str, month: str, publish_date: str = ""):
    """Генерує PDF таблицю результатів."""
    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm,  bottomMargin=1.5*cm,
    )

    # --- Стилі ---
    cell_s = ParagraphStyle("cell", fontName=FONT_REGULAR, fontSize=8, leading=10, wordWrap="CJK")
    pib_s  = ParagraphStyle("pib",  fontName=FONT_BOLD,    fontSize=9, leading=11, wordWrap="CJK")
    hdr_s  = ParagraphStyle("hdr",  fontName=FONT_BOLD,    fontSize=8, leading=10,
                             textColor=colors.white, alignment=1)
    title_s = ParagraphStyle("title", fontName=FONT_BOLD,  fontSize=14, leading=17,
                              textColor=colors.HexColor("#1a237e"), spaceAfter=4)
    instr_s = ParagraphStyle("instr", fontName=FONT_REGULAR, fontSize=9, leading=13,
                              textColor=colors.HexColor("#333333"))

    story = []

    # --- Назва ---
    story.append(Paragraph(f"Результати фестивалю TORONTO — {month}", title_s))

    # --- Інструкція ---
    instr = HEADER_TEXT
    if publish_date:
        instr += f"\nПублікація онлайн дипломів запланована на {publish_date} на сайті https://toronto.org.ua/"
    story.append(Paragraph(instr.replace("\n", "<br/>"), instr_s))
    story.append(Spacer(1, 0.3*cm))

    # --- Легенда кольорів ---
    def _badge(text, bg, fg):
        p = Paragraph(f"<b>{text}</b>",
                      ParagraphStyle("b", fontName=FONT_BOLD, fontSize=8,
                                     textColor=fg, leading=10))
        t = Table([[p]], colWidths=[2.5*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND",   (0,0),(-1,-1), bg),
            ("BOX",          (0,0),(-1,-1), 0.5, colors.HexColor("#AAAAAA")),
            ("LEFTPADDING",  (0,0),(-1,-1), 6),
            ("RIGHTPADDING", (0,0),(-1,-1), 6),
            ("TOPPADDING",   (0,0),(-1,-1), 2),
            ("BOTTOMPADDING",(0,0),(-1,-1), 2),
        ]))
        return t

    legend_items = [
        [_badge("Gran Pri",   ROW_COLORS["Gran Pri"],   LAU_TEXT_COLORS["Gran Pri"]),
         _badge("1st degree", ROW_COLORS["1st degree"], LAU_TEXT_COLORS["1st degree"]),
         _badge("2nd degree", ROW_COLORS["2nd degree"], LAU_TEXT_COLORS["2nd degree"]),
         _badge("3d degree",  ROW_COLORS["3d degree"],  LAU_TEXT_COLORS["3d degree"])],
    ]
    legend = Table(legend_items, colWidths=[2.7*cm]*4, hAlign="LEFT")
    legend.setStyle(TableStyle([
        ("LEFTPADDING",  (0,0),(-1,-1), 0),
        ("RIGHTPADDING", (0,0),(-1,-1), 6),
        ("TOPPADDING",   (0,0),(-1,-1), 0),
        ("BOTTOMPADDING",(0,0),(-1,-1), 0),
    ]))
    story.append(legend)
    story.append(Spacer(1, 0.25*cm))

    # --- Сортуємо за алфавітом ---
    sorted_rows = sorted(rows, key=lambda r: r["pib"].lower())

    # --- Таблиця ---
    col_widths = [1.3*cm, 5.5*cm, 3.5*cm, 5.3*cm, 2.8*cm]
    table_data = [[
        Paragraph("ID",                      hdr_s),
        Paragraph("ПІБ Учасника",            hdr_s),
        Paragraph("Номінація",               hdr_s),
        Paragraph("Назва або опис роботи",   hdr_s),
        Paragraph("Laureate",                hdr_s),
    ]]

    row_bg_styles  = []
    lau_txt_styles = []

    for i, r in enumerate(sorted_rows, start=1):
        lau   = r["laureate"]
        nazva = _clean_nazva(r.get("nazva", "") or "")
        row_bg  = ROW_COLORS.get(lau, colors.white)
        lau_fg  = LAU_TEXT_COLORS.get(lau, colors.black)
        row_bg_styles.append((i, row_bg))

        lau_para = Paragraph(
            f"<b>{lau}</b>",
            ParagraphStyle("lau_p", fontName=FONT_BOLD, fontSize=8,
                           leading=10, textColor=lau_fg),
        )
        table_data.append([
            Paragraph(str(r["id"]) if r["id"] else "—", cell_s),
            Paragraph(r["pib"],  pib_s),
            Paragraph(r["nom"],  cell_s),
            Paragraph(nazva,     cell_s),
            lau_para,
        ])

    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)

    ts = [
        ("BACKGROUND",    (0,0), (-1,0),  colors.HexColor("#1a237e")),
        ("TEXTCOLOR",     (0,0), (-1,0),  colors.white),
        ("FONTNAME",      (0,0), (-1,0),  FONT_BOLD),
        ("FONTSIZE",      (0,0), (-1,-1), 8),
        ("GRID",          (0,0), (-1,-1), 0.3, colors.HexColor("#CCCCCC")),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("ALIGN",         (0,0), (-1,0),  "CENTER"),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, colors.HexColor("#FAFAFA")]),
        ("LEFTPADDING",   (0,0), (-1,-1), 5),
        ("RIGHTPADDING",  (0,0), (-1,-1), 5),
        ("TOPPADDING",    (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]
    # Кольорові рядки по ступеню
    for row_idx, bg in row_bg_styles:
        ts.append(("BACKGROUND", (0,row_idx), (-1,row_idx), bg))

    tbl.setStyle(TableStyle(ts))
    story.append(tbl)

    doc.build(story)
    print(f"PDF збережено: {output_path}")


# ---------------------------------------------------------------------------
# Імпорт результатів з PDF або Excel попереднього конкурсу
# ---------------------------------------------------------------------------

# Стандартні назви колонок результатів (варіанти написання)
_COL_ALIASES = {
    "id":       ["id"],
    "pib":      ["піб учасника", "пib учасника", "учасник", "artist", "піб", "pib"],
    "nom":      ["номінація", "nomination"],
    "nazva":    ["назва або опис роботи", "назва або опис\nроботи", "назва роботи",
                 "назва композиції", "назва або опис", "назва"],
    "school":   ["назва навчального закладу", "навчальний заклад", "школа", "school", "organization"],
    "laureate": ["laureate", "лауреат", "ступінь", "degree"],
}

def _match_col(header: str) -> str | None:
    """Повертає ключ поля для заголовку колонки або None."""
    h = header.strip().lower().replace("\n", " ")
    for key, aliases in _COL_ALIASES.items():
        if any(h == a or h.startswith(a) for a in aliases):
            return key
    return None


def import_results_from_excel(path: str) -> tuple[list[dict], list[str]]:
    """
    Читає таблицю результатів з xlsx.
    Підтримує формат: ID | ПІБ Учасника | Номінація | Назва або опис роботи | Laureate
    Повертає (рядки, лог).
    """
    import openpyxl
    fname = os.path.basename(path)
    log   = [f"📂 Excel: {fname}"]
    wb    = openpyxl.load_workbook(path)
    rows  = []

    for sh in wb.sheetnames:
        ws = wb[sh]
        if not hasattr(ws, "iter_rows"):
            continue
        headers_raw = [str(c.value).strip() if c.value else "" for c in ws[1]]
        col_map = {}   # key -> column_index
        for i, h in enumerate(headers_raw):
            key = _match_col(h)
            if key and key not in col_map:
                col_map[key] = i

        log.append(f"  Аркуш '{sh}' | знайдено колонки: {list(col_map.keys())}")

        if "pib" not in col_map or "laureate" not in col_map:
            log.append(f"  🚫 Пропущено — немає обов'язкових колонок ПІБ / Laureate")
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            pib = row[col_map["pib"]] if "pib" in col_map else None
            if not pib or str(pib).strip().lower().startswith(
                    ("сума","гонорар","разом","total","підсумок","laureate","піб")):
                continue
            lau_raw = row[col_map["laureate"]] if "laureate" in col_map else None
            lau     = convert_laureate(lau_raw)
            school  = str(row[col_map["school"]]).strip() if "school" in col_map and row[col_map["school"]] else ""
            rows.append({
                "id":          str(row[col_map["id"]]).strip() if "id" in col_map and row[col_map["id"]] else "",
                "pib":         str(pib).strip(),
                "nom":         str(row[col_map["nom"]]).strip()   if "nom"   in col_map and row[col_map["nom"]]   else "",
                "vik":         "",
                "nazva":       str(row[col_map["nazva"]]).strip() if "nazva" in col_map and row[col_map["nazva"]] else "",
                "school":      school,
                "country":     detect_country(school),
                "laureate":    lau,
                "raw_laureate": str(lau_raw) if lau_raw is not None else "None",
                "comment":     "",
                "source":      fname,
            })

    log.append(f"  ✅ Імпортовано: {len(rows)} рядків")
    return rows, log


def import_results_from_pdf(path: str) -> tuple[list[dict], list[str]]:
    """
    Читає таблицю результатів з PDF (формат Toronto).
    Підтримує PDF згенеровані агентом результатів.
    Повертає (рядки, лог).
    """
    try:
        import pdfplumber
    except ImportError:
        return [], ["❌ pdfplumber не встановлений: pip install pdfplumber"]

    fname = os.path.basename(path)
    log   = [f"📂 PDF: {fname}"]
    rows  = []
    col_map = {}

    with pdfplumber.open(path) as pdf:
        log.append(f"  Сторінок: {len(pdf.pages)}")
        for page in pdf.pages:
            for table in page.extract_tables():
                if not table:
                    continue
                for row in table:
                    if not row:
                        continue
                    # Перший рядок з відомими заголовками → будуємо col_map
                    if not col_map:
                        candidate = {_match_col(str(c or "")): i
                                     for i, c in enumerate(row)
                                     if _match_col(str(c or ""))}
                        if "pib" in candidate and "laureate" in candidate:
                            col_map = candidate
                            log.append(f"  Знайдено заголовки: {list(col_map.keys())}")
                            continue
                        # Повторний заголовок на новій сторінці — пропускаємо
                        if col_map and _match_col(str(row[col_map.get("pib",1)] or "")) == "pib":
                            continue

                    if not col_map:
                        continue

                    pib = row[col_map["pib"]] if "pib" in col_map else None
                    if not pib:
                        continue
                    pib_s = str(pib).replace("\n"," ").strip()
                    if not pib_s or pib_s.lower() in ("піб учасника","artist","—",""):
                        continue

                    lau_raw = row[col_map["laureate"]] if "laureate" in col_map else None
                    lau_s   = str(lau_raw or "").replace("\n"," ").strip()
                    lau     = convert_laureate(lau_s) if lau_s else "1st degree"

                    rid = ""
                    if "id" in col_map:
                        rid_raw = str(row[col_map["id"]] or "").strip()
                        rid = rid_raw if rid_raw.isdigit() else ""

                    school = str(row[col_map["school"]] or "").replace("\n"," ").strip() if "school" in col_map else ""
                    rows.append({
                        "id":          rid,
                        "pib":         pib_s,
                        "nom":         str(row[col_map["nom"]] or "").replace("\n"," ").strip() if "nom"   in col_map else "",
                        "vik":         "",
                        "nazva":       str(row[col_map["nazva"]] or "").replace("\n"," ").strip() if "nazva" in col_map else "",
                        "school":      school,
                        "country":     detect_country(school),
                        "laureate":    lau,
                        "raw_laureate": lau_s or "None",
                        "comment":     "",
                        "source":      fname,
                    })

    log.append(f"  ✅ Імпортовано: {len(rows)} рядків")
    return rows, log


# ---------------------------------------------------------------------------
# Запис у Bitrix24
# ---------------------------------------------------------------------------

# UF_CRM_1690186104647 — Laureate (enumeration)
# UF_CRM_1702407283770 — Коментар Журі (string)
LAUREATE_ENUM_ID = {
    "Gran Pri":   184,
    "1st degree": 178,
    "2nd degree": 180,
    "3d degree":  182,
}

def write_to_bitrix(
    rows: list[dict],
    webhook_url: str,
    write_laureate: bool = True,
    write_comment: bool = True,
    progress_cb=None,          # callable(done, total, row, status) — для Streamlit
) -> dict:
    """
    Записує Laureate та/або Коментар Журі у Bitrix24 по ID угоди.
    Повертає {'ok': N, 'err': N, 'skip': N, 'errors': [(id, msg), ...]}.
    """
    webhook_url = webhook_url.rstrip("/")
    endpoint    = f"{webhook_url}/crm.deal.update.json"
    ok = err = skip = 0
    errors = []
    total  = len(rows)

    for i, r in enumerate(rows):
        rid = r.get("id")
        lau = r.get("laureate", "")
        comment = r.get("comment", "") or ""

        # Пропускаємо якщо немає ID
        if not rid:
            skip += 1
            if progress_cb:
                progress_cb(i + 1, total, r, "skip")
            continue
        try:
            rid_int = int(str(rid).strip())
        except (ValueError, TypeError):
            skip += 1
            if progress_cb:
                progress_cb(i + 1, total, r, "skip")
            continue

        # Формуємо fields
        fields = {}
        if write_laureate and lau:
            enum_id = LAUREATE_ENUM_ID.get(lau)
            if enum_id:
                fields["UF_CRM_1690186104647"] = enum_id
        if write_comment and comment:
            fields["UF_CRM_1702407283770"] = comment

        if not fields:
            skip += 1
            if progress_cb:
                progress_cb(i + 1, total, r, "skip")
            continue

        try:
            resp = requests.post(
                endpoint,
                json={"id": rid_int, "fields": fields},
                timeout=15,
            )
            data = resp.json() if resp.ok else {}
            if data.get("result"):
                ok += 1
                if progress_cb:
                    progress_cb(i + 1, total, r, "ok")
            else:
                msg = resp.text[:200]
                errors.append((rid_int, msg))
                err += 1
                if progress_cb:
                    progress_cb(i + 1, total, r, f"err:{msg}")
        except Exception as e:
            errors.append((rid_int, str(e)))
            err += 1
            if progress_cb:
                progress_cb(i + 1, total, r, f"err:{e}")

    return {"ok": ok, "err": err, "skip": skip, "errors": errors}


# ---------------------------------------------------------------------------
# Звіт
# ---------------------------------------------------------------------------
def print_report(rows: list[dict]):
    from collections import Counter
    cnt = Counter(r["laureate"] for r in rows)
    print("\n=== ЗВІТ ===")
    print(f"Всього учасників: {len(rows)}")
    for lau in ["Gran Pri", "1st degree", "2nd degree", "3d degree"]:
        print(f"  {lau}: {cnt.get(lau, 0)}")
    no_id = [r for r in rows if not r["id"]]
    if no_id:
        print(f"\nБез ID ({len(no_id)}):")
        for r in no_id:
            print(f"  {r['pib']}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Агент публікації результатів фестивалю TORONTO"
    )
    parser.add_argument("--folder",      required=True,
                        help="Папка з xlsx файлами оцінок від журі")
    parser.add_argument("--month",       required=True,
                        help='Місяць для заголовку PDF, наприклад "Квітень 2026"')
    parser.add_argument("--output",      default=None,
                        help="Шлях до вихідного PDF (за замовчуванням — поруч з --folder)")
    parser.add_argument("--publish-date", default="",
                        help='Дата публікації онлайн дипломів, наприклад "20 квітня"')
    parser.add_argument("--bitrix-url",  default="",
                        help="Webhook URL Bitrix24")
    parser.add_argument("--write-bitrix", action="store_true",
                        help="Записати Laureate у Bitrix24")
    args = parser.parse_args()

    folder = args.folder
    if not os.path.isdir(folder):
        print(f"ПОМИЛКА: папка не існує: {folder}")
        sys.exit(1)

    # Вихідний файл
    if args.output:
        output_pdf = args.output
    else:
        safe_month = re.sub(r'[^\w\s\-]', '', args.month).strip().upper().replace(" ", "_")
        output_pdf = os.path.join(
            os.path.dirname(folder),
            f"!!!РЕЗУЛЬТАТИ-{safe_month}.pdf"
        )

    print(f"Читаю файли журі з: {folder}")
    rows = read_all_jury(folder)
    if not rows:
        print("ПОМИЛКА: не знайдено жодного рядка з оцінками")
        sys.exit(1)

    print(f"\nЗнайдено учасників: {len(rows)}")
    print(f"Генерую PDF: {output_pdf}")
    build_pdf(rows, output_pdf, args.month, args.publish_date)

    if args.write_bitrix:
        if not args.bitrix_url:
            print("ПОМИЛКА: вкажіть --bitrix-url")
        else:
            print("\nЗаписую у Bitrix24...")
            write_to_bitrix(rows, args.bitrix_url)

    print_report(rows)
    print(f"\nГотово! PDF: {output_pdf}")


if __name__ == "__main__":
    import sys
    if sys.stdout.encoding != "utf-8":
        sys.stdout.reconfigure(encoding="utf-8")
    main()
